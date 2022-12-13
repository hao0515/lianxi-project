import requests
import json
from openpyxl import load_workbook


url = "https://pro-center-api.jinduo.com/open/api/accessory/product/list"                                            #将host与endpoint进行合并

data ={"page":1,"page_size":500,"queryAccessoryProductIds":[503,508,510,505,507,509,512,1,2,3,4,5,6,7,8,9,10,11,12,13,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,149,159,171,181,191,203,465,466,467,468,469,492,493,494,501,502,504,506,511,523,524,525,526,59,60,61,62,63,67,68,69,70,71,75,76,77,78,79,83,84,85,86,87,91,92,93,94,95,98,99,100,101,102,103,106,107,108,109,110,111,114,115,116,117,118,119,122,123,124,125,126,127,128,130,131,132,133,134,135,138,139,140,141,142,143,146,147,148,150,151,154,155,156,157,158,162,163,164,165,166,167,170,172,173,174,175,178,179,180,182,183,186,187,188,189,190,194,204,473,476,479,480,481,484,495,496,497,498,499,257,513,258,259,266,267,272,273,274,275,280,281,306,307,57,58,64,65,66,72,73,74,80,81,82,88,89,90,96,97,104,105,112,113,120,121,129,136,137,144,145,152,153,160,161,168,169,176,177,184,185,192,193,205,206,207,208,209,210,211,212,213,214,215,471,216,472,217,474,475,477,478,482,483,485,486,487,488,489,490,243,500,249,250,251,256,268,269,270,276,277,278,279,308,309,310,322,323,324,325,326,327,328,329,330,331,332,333,334,348,350,351,360,361,362,363,366,377,378,379,380,388,389,390,391,392,393,394,395,396,397,398,399,432,463,464,470,244,245,246,247,248,252,253,254,255,514,515,260,516,261,517,262,518,263,519,264,265,521,522,14,15,271,16,17,18,19,20,21,22,23,24,25,26,282,27,283,28,284,285,286,287,288,289,290,291,292,293,294,295,296,297,298,299,300,301,302,303,304,305,311,312,313,314,315,316,317,318,319,320,321,335,336,337,338,339,340,341,342,343,344,345,346,347,349,352,353,354,355,356,357,358,359,364,365,367,368,369,370,371,372,373,374,375,376,381,382,383,384,385,386,387,400,401,402,403,404,405,406,407,408,409,410,411,412,413,414,415,416,417,418,419,420,421,422,423,424,425,426,427,428,429,430,431,433,434,435,436,437,438,439,440,441,442,443,444,445,446,447,448,449,450,195,451,196,452,197,453,198,454,199,455,200,456,201,457,202,458,459,460,461,462,218,219,220,221,222,223,224,225,226,227,228,229,230,231,232,233,234,235,491,236,237,238,239,240,241,242],"plan_status":"主推产品","hasParamOptions":1,"brand_id":6}

                         #将data（字典）表单形式数据进行序列化
r = requests.post(url,json = data)                                         #不用序列化，直接转型，使用方便

                                                         #打印响应的响应体信息
resp = r.json()['data']['data'][0]['accessory_products'][0]['model_num']
#print(resp)                                                        #将响应的信息已json串的形式展示出来
i = 0
while (i < 345):
    resp = r.json()['data']['data'][i]['accessory_products'][0]['model_num']
    print(f"'{resp}',")

    wb = load_workbook(r'E:\automation\pythonProject2\OSS列表型号.xlsx')
    sheet = wb.active
    sheet['a1'] = '型号'
    sheet.cell(row = 2, column = 2).value = resp
    i = i + 1
    wb.save(r'E:\automation\pythonProject2\OSS列表型号.xlsx')
    print('数据写入成功！')